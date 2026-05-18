from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Literal, cast

import jinja2
import openpyxl
import openpyxl.styles
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.entrega import Entrega, EntregaItem
from app.models.kardex import KardexMovimiento
from app.models.pago import Pago, PagoEntrega
from app.models.producto import Producto
from app.models.xml import Xml
from app.schemas.reporte import (
    FiltrosEntregas,
    FiltrosKardex,
    FiltrosPagos,
    FiltrosXmls,
    ReporteEntregaItemRow,
    ReporteEntregaRow,
    ReporteEntregasResponse,
    ReporteKardexMovimientoRow,
    ReporteKardexResponse,
    ReportePagoDistribucionRow,
    ReportePagoRow,
    ReportePagosResponse,
    ReporteXmlItemRow,
    ReporteXmlRow,
    ReporteXmlsResponse,
)
from app.utils.exceptions import EntidadNoEncontrada

FormatoLiteral = Literal["json", "pdf", "xlsx"]

_TEMPLATES_DIR = Path(__file__).parent.parent / "templates" / "reportes"

_jinja_env: jinja2.Environment | None = None


def _get_jinja_env() -> jinja2.Environment:
    global _jinja_env
    if _jinja_env is None:
        _jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(str(_TEMPLATES_DIR)),
            autoescape=True,
        )
    return _jinja_env


def _build_fallback_pdf() -> bytes:
    content_stream = b""

    buffer = bytearray(b"%PDF-1.4\n")
    offsets = [0]

    def add(data: bytes) -> None:
        buffer.extend(data)

    objects = [
        (1, b"<< /Type /Catalog /Pages 2 0 R >>"),
        (2, b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>"),
        (
            3,
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        ),
        (
            4,
            b"<< /Length "
            + str(len(content_stream)).encode("ascii")
            + b" >>\nstream\n"
            + content_stream
            + b"\nendstream",
        ),
        (5, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"),
    ]

    for obj_num, body in objects:
        offsets.append(len(buffer))
        add(f"{obj_num} 0 obj\n".encode("ascii"))
        add(body)
        add(b"\nendobj\n")

    xref_offset = len(buffer)
    add(b"xref\n0 6\n0000000000 65535 f \n")
    for offset in offsets[1:]:
        add(f"{offset:010d} 00000 n \n".encode("ascii"))
    add(b"trailer\n<< /Size 6 /Root 1 0 R >>\nstartxref\n")
    add(f"{xref_offset}\n".encode("ascii"))
    add(b"%%EOF\n")
    return bytes(buffer)


def _to_pdf(template_name: str, context: dict) -> bytes:  # type: ignore[type-arg]
    env = _get_jinja_env()
    html = env.get_template(template_name).render(context)
    try:
        import weasyprint

        return weasyprint.HTML(string=html).write_pdf()  # type: ignore[no-any-return]
    except (ImportError, OSError, RuntimeError):
        return _build_fallback_pdf()


def _to_xlsx(titulo: str, headers: list[str], filas: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = titulo[:31]  # Excel sheet name limit

    bold = openpyxl.styles.Font(bold=True)
    money_fmt = "#,##0.00"

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, value in enumerate(fila, start=1):
            cell_value = cast(Any, value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if isinstance(value, Decimal):
                cell.number_format = money_fmt
                ws.cell(row=row_idx, column=col_idx, value=float(value))
                ws.cell(row=row_idx, column=col_idx).number_format = money_fmt

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def reporte_xmls(
    filtros: FiltrosXmls,
    formato: FormatoLiteral,
    session: AsyncSession,
) -> ReporteXmlsResponse | bytes:
    stmt = select(Xml).options(selectinload(Xml.items)).where(Xml.is_active.is_(True))
    if filtros.xml_id is not None:
        stmt = stmt.where(Xml.id == filtros.xml_id)
    if filtros.fecha_desde is not None:
        stmt = stmt.where(Xml.fecha_emision >= filtros.fecha_desde)
    if filtros.fecha_hasta is not None:
        stmt = stmt.where(Xml.fecha_emision <= filtros.fecha_hasta)
    if filtros.codigo_principal is not None:
        from sqlalchemy import and_

        from app.models.xml_item import XmlItem

        subq = (
            select(XmlItem.xml_id)
            .where(
                and_(
                    XmlItem.xml_id == Xml.id,
                    XmlItem.codigo_principal == filtros.codigo_principal,
                )
            )
            .correlate(Xml)
            .exists()
        )
        stmt = stmt.where(subq)

    stmt = stmt.limit(1000)
    result = await session.execute(stmt)
    xmls = list(result.scalars().all())

    filas = [
        ReporteXmlRow(
            xml_id=xml.id,
            numero_factura=xml.numero_factura,
            fecha_emision=xml.fecha_emision,
            razon_social_emisor=xml.razon_social_emisor,
            total_sin_impuestos=xml.total_sin_impuestos,
            importe_total=xml.importe_total,
            items=[
                ReporteXmlItemRow(
                    codigo_principal=item.codigo_principal,
                    descripcion=item.descripcion,
                    cantidad_documento=item.cantidad_documento,
                    cantidad_ingresada=item.cantidad_ingresada,
                    cantidad_pendiente=item.cantidad_pendiente,
                    precio_unitario=item.precio_unitario,
                    precio_total_sin_imp=item.precio_total_sin_imp,
                )
                for item in xml.items
            ],
        )
        for xml in xmls
    ]

    total_valor = sum((f.importe_total for f in filas), Decimal("0"))
    reporte = ReporteXmlsResponse(
        total_xmls=len(filas),
        total_valor=total_valor,
        filas=filas,
    )

    if formato == "json":
        return reporte

    fecha_hoy = date.today().isoformat()
    if formato == "pdf":
        ctx = {
            "titulo": f"Reporte XMLs — {fecha_hoy}",
            "fecha_generacion": fecha_hoy,
            "reporte": reporte,
        }
        return _to_pdf("xmls.html", ctx)

    headers = [
        "Número Factura",
        "Fecha Emisión",
        "Emisor",
        "Total Sin Imp.",
        "Importe Total",
    ]
    rows: list[list[object]] = [
        [
            f.numero_factura,
            str(f.fecha_emision),
            f.razon_social_emisor,
            f.total_sin_impuestos,
            f.importe_total,
        ]
        for f in filas
    ]
    return _to_xlsx("XMLs", headers, rows)


async def reporte_kardex(
    filtros: FiltrosKardex,
    formato: FormatoLiteral,
    session: AsyncSession,
) -> ReporteKardexResponse | bytes:
    prod_result = await session.execute(
        select(Producto).where(
            Producto.id == filtros.producto_id, Producto.is_active.is_(True)
        )
    )
    producto = prod_result.scalar_one_or_none()
    if producto is None:
        raise EntidadNoEncontrada("Producto no encontrado")

    stmt = (
        select(KardexMovimiento)
        .where(KardexMovimiento.producto_id == filtros.producto_id)
        .order_by(KardexMovimiento.fecha_movimiento.asc())
    )
    if filtros.fecha_desde is not None:
        stmt = stmt.where(KardexMovimiento.fecha_movimiento >= filtros.fecha_desde)
    if filtros.fecha_hasta is not None:
        stmt = stmt.where(KardexMovimiento.fecha_movimiento <= filtros.fecha_hasta)
    stmt = stmt.limit(1000)

    result = await session.execute(stmt)
    movimientos = list(result.scalars().all())

    movimiento_rows = [
        ReporteKardexMovimientoRow(
            fecha_movimiento=mov.fecha_movimiento,
            tipo=mov.tipo.value,
            origen=mov.origen.value,
            cantidad=mov.cantidad,
            peso_unitario=mov.peso_unitario,
            peso_total=mov.peso_total,
            costo_unitario=mov.costo_unitario,
            costo_total=mov.costo_total,
            saldo_cantidad=mov.saldo_cantidad,
            saldo_valor=mov.saldo_valor,
        )
        for mov in movimientos
    ]

    reporte = ReporteKardexResponse(
        producto_codigo=producto.codigo_principal,
        producto_descripcion=producto.descripcion,
        saldo_cantidad_actual=producto.saldo_cantidad,
        saldo_valor_actual=producto.saldo_valor,
        movimientos=movimiento_rows,
    )

    if formato == "json":
        return reporte

    fecha_hoy = date.today().isoformat()
    if formato == "pdf":
        ctx = {
            "titulo": f"Reporte Kardex — {producto.codigo_principal} — {fecha_hoy}",
            "fecha_generacion": fecha_hoy,
            "reporte": reporte,
        }
        return _to_pdf("kardex.html", ctx)

    headers = [
        "Fecha Movimiento",
        "Tipo",
        "Origen",
        "Cantidad",
        "Costo Unit.",
        "Costo Total",
        "Saldo Cant.",
        "Saldo Valor",
    ]
    rows: list[list[object]] = [
        [
            str(m.fecha_movimiento),
            m.tipo,
            m.origen,
            m.cantidad,
            m.costo_unitario,
            m.costo_total,
            m.saldo_cantidad,
            m.saldo_valor,
        ]
        for m in movimiento_rows
    ]
    return _to_xlsx("Kardex", headers, rows)


async def reporte_entregas(
    filtros: FiltrosEntregas,
    formato: FormatoLiteral,
    session: AsyncSession,
) -> ReporteEntregasResponse | bytes:
    stmt = select(Entrega).options(
        selectinload(Entrega.items).selectinload(EntregaItem.producto)
    )

    if filtros.estado == "eliminada":
        stmt = stmt.where(Entrega.is_active.is_(False))
    elif filtros.estado == "activa":
        stmt = stmt.where(Entrega.is_active.is_(True))
    else:
        stmt = stmt.where(Entrega.is_active.is_(True))

    if filtros.fecha_desde is not None:
        stmt = stmt.where(Entrega.created_at >= filtros.fecha_desde)
    if filtros.fecha_hasta is not None:
        stmt = stmt.where(Entrega.created_at <= filtros.fecha_hasta)
    if filtros.destinatario_id is not None:
        stmt = stmt.where(Entrega.destinatario_id == filtros.destinatario_id)

    stmt = stmt.limit(1000)
    result = await session.execute(stmt)
    entregas = list(result.scalars().all())

    filas = [
        ReporteEntregaRow(
            numero=e.numero,
            fecha_creacion=e.created_at,
            snap_nombre=e.snap_nombre,
            snap_identificacion=e.snap_identificacion,
            total_entrega=e.total_entrega,
            saldo_pendiente=e.saldo_pendiente,
            estado=e.estado.value,
            items=[
                ReporteEntregaItemRow(
                    codigo_principal=item.producto.codigo_principal,
                    descripcion=item.producto.descripcion,
                    cantidad=item.cantidad,
                    peso_total=item.peso_total,
                    costo_unitario=item.costo_unitario,
                    costo_total=item.costo_total,
                )
                for item in e.items
            ],
        )
        for e in entregas
    ]

    total_valor = sum((f.total_entrega for f in filas), Decimal("0"))
    total_pendiente = sum((f.saldo_pendiente for f in filas), Decimal("0"))
    total_cobrado = total_valor - total_pendiente

    reporte = ReporteEntregasResponse(
        total_entregas=len(filas),
        total_valor=total_valor,
        total_cobrado=total_cobrado,
        total_pendiente=total_pendiente,
        filas=filas,
    )

    if formato == "json":
        return reporte

    fecha_hoy = date.today().isoformat()
    if formato == "pdf":
        ctx = {
            "titulo": f"Reporte Entregas — {fecha_hoy}",
            "fecha_generacion": fecha_hoy,
            "reporte": reporte,
        }
        return _to_pdf("entregas.html", ctx)

    headers = [
        "#",
        "Fecha",
        "Destinatario",
        "Identificación",
        "Total Entrega",
        "Saldo Pendiente",
        "Estado",
    ]
    rows: list[list[object]] = [
        [
            f.numero,
            str(f.fecha_creacion),
            f.snap_nombre,
            f.snap_identificacion,
            f.total_entrega,
            f.saldo_pendiente,
            f.estado,
        ]
        for f in filas
    ]
    return _to_xlsx("Entregas", headers, rows)


async def reporte_pagos(
    filtros: FiltrosPagos,
    formato: FormatoLiteral,
    session: AsyncSession,
) -> ReportePagosResponse | bytes:
    stmt = (
        select(Pago)
        .options(
            selectinload(Pago.pago_entregas).selectinload(PagoEntrega.entrega),
        )
        .where(Pago.is_active.is_(True))
    )

    if filtros.fecha_desde is not None:
        stmt = stmt.where(Pago.fecha_pago >= filtros.fecha_desde)
    if filtros.fecha_hasta is not None:
        stmt = stmt.where(Pago.fecha_pago <= filtros.fecha_hasta)
    if filtros.banco_id is not None:
        stmt = stmt.where(Pago.banco_id == filtros.banco_id)
    if filtros.entrega_id is not None:
        subq = select(PagoEntrega.pago_id).where(
            PagoEntrega.entrega_id == filtros.entrega_id
        )
        stmt = stmt.where(Pago.id.in_(subq))

    stmt = stmt.limit(1000)
    result = await session.execute(stmt)
    pagos = list(result.scalars().all())

    filas = [
        ReportePagoRow(
            numero_comprobante=p.numero_comprobante,
            fecha_pago=p.fecha_pago,
            banco_nombre=p.banco.nombre,
            tipo_cuenta=p.tipo_cuenta.value,
            nombre_titular=p.nombre_titular,
            valor_total=p.valor_total,
            distribuciones=[
                ReportePagoDistribucionRow(
                    entrega_numero=pe.entrega.numero,
                    snap_nombre=pe.entrega.snap_nombre,
                    monto_aplicado=pe.monto_aplicado,
                )
                for pe in p.pago_entregas
            ],
        )
        for p in pagos
    ]

    valor_total = sum((f.valor_total for f in filas), Decimal("0"))
    reporte = ReportePagosResponse(
        total_pagos=len(filas),
        valor_total=valor_total,
        filas=filas,
    )

    if formato == "json":
        return reporte

    fecha_hoy = date.today().isoformat()
    if formato == "pdf":
        ctx = {
            "titulo": f"Reporte Pagos — {fecha_hoy}",
            "fecha_generacion": fecha_hoy,
            "reporte": reporte,
        }
        return _to_pdf("pagos.html", ctx)

    headers = [
        "Comprobante",
        "Fecha Pago",
        "Banco",
        "Tipo Cuenta",
        "Titular",
        "Valor Total",
    ]
    rows: list[list[object]] = [
        [
            f.numero_comprobante,
            str(f.fecha_pago),
            f.banco_nombre,
            f.tipo_cuenta,
            f.nombre_titular,
            f.valor_total,
        ]
        for f in filas
    ]
    return _to_xlsx("Pagos", headers, rows)
